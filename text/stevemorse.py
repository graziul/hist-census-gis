from openpyxl import Workbook
import urllib
import pandas as pd
import re
import pickle

# Use one year's CityInfo file to get list of states
file_path = '/home/s4-data/LatestCities'

# Scrape Steve Morse street-ed data from website
def scrape_sm_st_ed(file_path, decades=[1900,1910,1930,1940]):

	city_info_file = file_path + '/CityInfo.csv' 
	city_info_df = pd.read_csv(city_info_file)

	state_list = city_info_df['state_abbr'].str.lower().unique().tolist()
	city_list = city_info_df['city_name'].tolist()

	city_state_iterator = zip(city_info_df['city_name'].tolist(),city_info_df['state_abbr'].tolist())

	# Ignore unicode characters
	def ignore_unicode(chars):
		chars = unicode(chars,errors='ignore')
		return chars

	# Standardize Steve Morse street names
	def sm_standardize(st) :
		orig_st = st
		st = re.sub(r" [Ee][Xx][Tt][Ee]?[Nn]?[Dd]?[Ee]?[Dd]?$","",st)
		DIR = re.search(r" ([NSEW ]+)$",st)
		st = re.sub(r" ([NSEW ]+)$","",st)
		if(DIR) :
			DIR = DIR.group(1)
			DIR = re.sub(" ","",DIR)
		else :
			DIR = ""

		TYPE = re.search(r' (St|Ave?|Blvd|Pl|Dr|Drive|Rd|Road|Ct|Railway|CityLimits|Hwy|Fwy|Pkwy|Cir|Ter|La|Ln|Way|Trail|Sq|All?e?y?|Bridge|Bridgeway|Walk|Crescent|Creek|River|Line|Plaza|Esplanade|[Cc]emetery|Viaduct|Trafficway|Trfy|Turnpike)$',st)
		
		if(TYPE) :
			st = re.sub(TYPE.group(0),"",st)
			TYPE = TYPE.group(1)
			if(TYPE=="Av") :
				TYPE = "Ave"
			if(TYPE=="Rd") :
				TYPE = "Road"
			if(TYPE=="Dr") :
				TYPE = "Drive"
			if(TYPE=="La") :
				TYPE = "Ln"            
			if(re.match("All?e?y?",TYPE)) :
				TYPE = "Aly"
		else :
			if re.search("[Cc]ity [Ll]imits|[Rr]ailroad [Tt]racks",orig_st) :
				TYPE = ""
			else :
				TYPE = "St"
		
		NAME = st
		st = (DIR+" "+NAME+" "+TYPE).strip()
		#print(orig_st)
		#print("changed to "+st)
		return [st,DIR,NAME,TYPE]
	
	# Download and save Steve Morse web abbreviations
	def download_sm_web_abbr(decades):

		# Get the city_state abbreviation for Steve Morse    
		def get_sm_web_abbr(decade):

			for city_state in city_state_iterator:

				city_name = city_state[0]
				state_abbr = city_state[1]

				sm_web_abbr_dict[decade][state_abbr][city_name] = []

				url = "http://stevemorse.org/census/%sstates/%s.htm" % (str(decade),state_abbr) 
				url_handle = urllib.urlopen(url)
				sourcetext = url_handle.readlines()
				url_handle.close()

				#Change the parsing strings if necessary (CG: Doesn't seem necessary yet)
				start_num = "value=" 
				end_num = ">"
				end_name = "</option>"

				for line in sourcetext:
					#If line has "value=" and "</option>" in it, do stuff
					if start_num in line and end_name in line:
						#Ignore header line
						if "Select City" in line:
							continue
						if "Other" in line:
							continue
						else:
							city = line[line.find(end_num)+1:line.find("(")].rstrip().replace('.','')
							if city_name in city:
								city_abbr = line[line.find(start_num) + 7 : line.find(end_num)-1].lower()
								city_abbr = re.sub(r'\$[0-9]|\*[0-9]','',city_abbr)
								sm_web_abbr = city_abbr + state_abbr
								if sm_web_abbr not in sm_web_abbr_dict[decade][state_abbr][city_name]:
									sm_web_abbr_dict[decade][state_abbr][city_name].append(sm_web_abbr)

		sm_web_abbr_dict = {}
		for decade in decades:
			sm_web_abbr_dict[decade] = {}
			for state_abbr in state_list:
				sm_web_abbr_dict[decade][state_abbr] = {}
			get_sm_web_abbr(decade)
		pickle.dump(sm_web_abbr_dict,open(file_path + '/sm_web_abbr.pickle','wb'))

	# Download and save Steve Morse street-ed information
	def download_sm_st_ed(decades):

		# Get Steve Morse street-ed information for a single decade
		def get_sm_st_ed(decade, sm_st_ed_dict):
				
			for city_state in city_state_iterator:

				city_name = city_state[0]
				state_abbr = city_state[1]
		  
				sm_st_ed_dict_city = {}

				try:
					sm_web_abbr = sm_web_abbr_dict[decade][state_abbr][city_name]
					if city_name == 'Springfield' and state_abbr = 'MA':
						sm_web_abbr = ['spma'] # Algorithm picks up "West Springfield" as well
				except:
					continue

				for i in sm_web_abbr:

					url = "http://www.stevemorse.org/census/%scities/%s.htm" % (str(decade),i.lower()) 

					url_handle = urllib.urlopen(url)
					sourcetext = url_handle.readlines()
					url_handle.close()

					#Change the parsing strings if necessary (CG: Doesn't seem necessary yet)
					start_num = "value=" 
					end_num = ">"
					end_name = "</option>"

					for line in sourcetext:
						#If line has "value=" and "</option>" in it, do stuff
						if start_num in line and end_name in line:
							#Ignore header line
							if "Select Street" in line:
								col_line = ""
							else:
								#Extract street name between ">" and "</option>""
								st = line[line.find(end_num)+1:line.find(end_name)]
								st = sm_standardize(st)
								#Initialize dictionary for NAME if it doesn't exist already
								NAME = st[2]
								sm_st_ed_dict_city[NAME] = sm_st_ed_dict_city.setdefault(NAME, {})
								#Extract EDs as one long string
								st_str = line[line.find(start_num) + 7 : line.find(end_num)-1]
								#Check for street name collisions (e.g. "3rd Pl" and "3rd Pl ext")
								if st[0] in sm_st_ed_dict_city[NAME].keys():
									for i in list(st_str.split(",")):
										if i not in sm_st_ed_dict_city[NAME][st[0]]:
											sm_st_ed_dict_city[NAME][st[0]].append(i)
								else:    
									#Split string of EDs into a list and assign to street name in dictionary
									sm_st_ed_dict_city[NAME].update({st[0]:list(st_str.split(","))})
				sm_st_ed_dict[decade][city_state] = sm_st_ed_dict_city

		# Run get_sm_st_ed for all decades, save results
		sm_st_ed_dict ={}
		for decade in decades:
			sm_st_ed_dict[decade] = {}
			for city_state in city_state_iterator:
				sm_st_ed_dict[decade][city_state] = {}
			get_sm_st_ed(decade, sm_st_ed_dict)
			pickle.dump(sm_st_ed_dict[decade],open(file_path + '/%s/sm_st_ed_dict%s.pickle' % (str(decade),str(decade)),'wb'))
		pickle.dump(sm_st_ed_dict,open(file_path + '/sm_st_ed_dict.pickle' % (str(decade),str(decade)),'wb'))

	# Output lists for use by students
	def write_sm_lists(decades):

		# Create usable ST-ED and ED-ST lists
		def create_usable_list(sm_st_ed_dict):
			sm_ed_st_dict = {}
			for city_state in city_state_iterator:
				sm_ed_st_dict[city_state] = {}
				sm_st_ed_dict_nested = sm_st_ed_dict[city_state] 
				#Flatten dictionary
				temp = {k:v for d in [v for k,v in sm_st_ed_dict_nested.items()] for k,v in d.items()}
				#Initialize a list of street names without an ED in Steve Morse
				sm_ed_st_dict[city_state][''] = []
				for st, eds in temp.items():
					#If street name has no associated EDs (i.e. street name not found in Steve Morse) 
					#then add to dictionary entry for no ED
					if eds is None:
						sm_ed_st_dict[city_state][''].append(st)
					else:
						#For every ED associated with a street name...
						for ed in eds:
							#Initalize an empty list if ED has no list of street names yet
							sm_ed_st_dict[city_state][ed] = sm_ed_st_dict[city_state].setdefault(ed, [])
							#Add street name to the list of streets
							sm_ed_st_dict[city_state][ed].append(st)
			return sm_ed_st_dict

		# Output list usable by students
		def output_usable_list(sm_dict, decade):

			sm_dict = sm_st_ed_dict[decade]

			for city_state in city_state_iterator:

				c_s = city_state[0] + city_state[1].upper()
				file_name = file_path + '/' + str(decade) + '/sm_lists/' + c_s + '_SM.xlsx'
				wb = Workbook()

				temp = sm_dict[city_state]
				sm_st_ed_dict_city = {k:v for d in [v for k,v in temp.items()] for k,v in d.items()}

				#Save street-to-ED sheet
				ws_st = wb.active
				ws_st.title = '%s Street to ED' % (str(decade)) 
				ws_st.append([str(decade) + ' Street',str(decade) + ' EDs'])
				dictlist = []
				keylist = sm_st_ed_dict_city.keys()
				keylist.sort()
				for k in keylist:
					t = [i.replace("'","") for i in sm_st_ed_dict_city[k]]
					ws_st.append([ignore_unicode(k)]+t)        

				#Save ED-to-street sheet
				ws_ed = wb.create_sheet('%s ED to Street' % (str(decade)))
				ws_ed.append([str(decade) + ' ED',str(decade) + ' Streets'])

				sm_ed_st_dict_city = {}
				for st, eds in sm_st_ed_dict_city.items():
					for ed in eds:
						#Initalize an empty list if ED has no list of street names yet
						sm_ed_st_dict_city[ed] = sm_ed_st_dict_city.setdefault(ed, [])
						#Add street name to the list of streets
						sm_ed_st_dict_city[ed].append(st)

				dictlist = []
				keylist = sm_ed_st_dict_city.keys()
				keylist.sort()
				for k in keylist:
					t = [ignore_unicode(i).replace("'","") for i in sm_ed_st_dict_city[k]]
					ws_ed.append([k]+t)       

				#If 1930, also include list of 1940 SM streets
				if decade == 1930:
					ws_st1940 = wb.create_sheet('1940 Streets')
					ws_st1940.append(['1940 Streets'])
					streetlist = sm_st_ed_dict[1940][city_state].keys()
					streetlist.sort()
					for i in streetlist:
						ws_st1940.append([i])

				wb.save(file_name)

		sm_st_ed_dict = pickle.load(open(file_path + '/sm_st_ed_dict.pickle','rb'))
		for decade in decades:
			sm_ed_st_dict = create_usable_list(sm_st_ed_dict[decade])
			output_usable_list(sm_ed_st_dict, decade)

	download_sm_web_abbr(decades)
	download_sm_web_abbr(decades)
	write_sm_lists(decades)

# Load Steve Morse st_ed data
def load_steve_morse(city_name, state_abbr, decade, file_path):

	if city_name == 'Richmond' and state_abbr == 'NY':
		city_name = 'Staten Island'
	#NOTE: This dictionary must be built independently of this script
	sm_st_ed_dict_file = pickle.load(open(file_path + '/%s/sm_st_ed_dict%s.pickle' % (str(decade), str(decade)), 'rb'))
	sm_st_ed_dict_nested = sm_st_ed_dict_file[(city_name, '%s' % (state_abbr))]

	#Flatten dictionary
	temp = {k:v for d in [v for k, v in sm_st_ed_dict_nested.items()] for k, v in d.items()}

	#Capture all Steve Morse streets in one list
	sm_all_streets = temp.keys()

	#
	# Build a Steve Morse (sm) ED-to-Street (ed_st) dictionary (dict)
	#

	sm_ed_st_dict = {}
	#Initialize a list of street names without an ED in Steve Morse
	sm_ed_st_dict[''] = []
	for st, eds in temp.items():
		#If street name has no associated EDs (i.e. street name not found in Steve Morse) 
		#then add to dictionary entry for no ED
		if eds is None:
			sm_ed_st_dict[''].append(st)
		else:
			#For every ED associated with a street name...
			for ed in eds:
				#Initalize an empty list if ED has no list of street names yet
				sm_ed_st_dict[ed] = sm_ed_st_dict.setdefault(ed, [])
				#Add street name to the list of streets
				sm_ed_st_dict[ed].append(st)

	return sm_all_streets, sm_st_ed_dict_nested, sm_ed_st_dict
