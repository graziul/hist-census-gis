from openpyxl import Workbook
from histcensusgis.text.standardize import *
import urllib
import pandas as pd
import re
import pickle
import os
import histcensusgis
import time
import math
import geopandas as gpd

# Use one year's CityInfo file to get list of states
file_path = '/home/s4-data/LatestCities'
package_path = os.path.dirname(histcensusgis.__file__)

# Ignore unicode characters
def ignore_unicode(chars):
	chars = unicode(chars,errors='ignore')
	return chars

# Create ED-street dictionary from street-ED dictionary 
def create_ed_st_dict(sm_st_ed_dict):
	sm_ed_st_dict = {}
	for i, city_state in city_state_iterator.iterrows():

		city_name = city_state[0]
		state_abbr = city_state[1]

		city_state = list(city_state)
		city_state[0] = city_state[0].replace(' ','')
		city_state = tuple(city_state)
		sm_ed_st_dict[city_state] = {}
		sm_st_ed_dict_nested = sm_st_ed_dict[city_state] 
		#Flatten dictionary
		temp = {k:v for d in [v for k,v in sm_st_ed_dict_nested.items()] for k,v in d.items()}
		#Initialize a list of street names without an ED in Steve Morse
		sm_ed_st_dict[city_state][''] = []
		for st, eds in temp.items():
			#If street name has no associated EDs (i.e. street name not found in Steve Morse) 
			#then add to dictionary entry for no ED
			if eds is None:
				sm_ed_st_dict[city_state][''].append(st)
			else:
				#For every ED associated with a street name...
				for ed in eds:
					#Initalize an empty list if ED has no list of street names yet
					sm_ed_st_dict[city_state][ed] = sm_ed_st_dict[city_state].setdefault(ed, [])
					#Add street name to the list of streets
					sm_ed_st_dict[city_state][ed].append(st)
	return sm_ed_st_dict

# Get the city_state abbreviation for Steve Morse    
def get_sm_web_abbr(decade,sm_web_abbr_dict):

	for i, city_state in city_state_iterator.iterrows():

		city_name = city_state[0]
		state_abbr = city_state[1]

		sm_web_abbr_dict[decade][state_abbr][city_name.replace(' ','')] = []

		url = "http://stevemorse.org/census/%sstates/%s.htm" % (str(decade),state_abbr) 
		url_handle = urllib.urlopen(url)
		sourcetext = url_handle.readlines()
		url_handle.close()

		#Change the parsing strings if necessary (CG: Doesn't seem necessary yet)
		start_num = "value=" 
		end_num = ">"
		alt_end_num = " selected>" # AFAIK, this applies only to DC '40
		end_name = "</option>"

		for line in sourcetext:
			#If line has "value=" and "</option>" in it, do stuff
			if start_num in line and end_name in line:
				#Ignore header line
				if "Select City" in line:
					continue
				if "Other" in line:
					continue
				else:
					city = line[line.find(end_num)+1:line.find("(")].rstrip().replace('.','')
					if city_name in city:
						if alt_end_num in line :
							city_abbr = line[line.find(start_num) + 7 : line.find(alt_end_num)-1].lower()
						else :
							city_abbr = line[line.find(start_num) + 7 : line.find(end_num)-1].lower()
						city_abbr = re.sub(r'\$[0-9]|\*[0-9]','',city_abbr)
						sm_web_abbr = city_abbr + state_abbr
						if sm_web_abbr not in sm_web_abbr_dict[decade][state_abbr][city_name.replace(' ','')]:
							sm_web_abbr_dict[decade][state_abbr][city_name.replace(' ','')].append(sm_web_abbr)
		time.sleep(1)

# Get Steve Morse street-ed information for a single decade
def get_sm_st_ed(decade, sm_st_ed_dict, package_path):
	sm_web_abbr_dict = pickle.load(open(package_path + '/text/sm_web_abbr.pickle','rb'))
	no_data = 0
	for i, city_state in city_state_iterator.iterrows():
		city_name = city_state[0]
		state_abbr = city_state[1]
		city_state = list(city_state)
		city_state[0] = city_state[0].replace(' ','')
		city_state = tuple(city_state)
		sm_st_ed_dict_city = {}
		try:
			sm_web_abbr = sm_web_abbr_dict[decade][state_abbr][city_name.replace(' ','')]
			if city_name == 'Springfield' and state_abbr == 'MA':
				sm_web_abbr = ['spma'] # Algorithm picks up "West Springfield" as well
			if sm_web_abbr == []:
				no_data+=1
				continue
		except:
			continue
		for i in sm_web_abbr:
			url = "http://www.stevemorse.org/census/%scities/%s.htm" % (str(decade),i.lower()) 
			url_handle = urllib.urlopen(url)
			sourcetext = url_handle.readlines()
			url_handle.close()
			#Change the parsing strings if necessary (CG: Doesn't seem necessary yet)
			start_num = "value=" 
			end_num = ">"
			end_name = "</option>"
			for line in sourcetext:
				#If line has "value=" and "</option>" in it, do stuff
				if start_num in line and end_name in line:
					#Ignore header line
					if "Select Street" in line:
						col_line = ""
					else:
						#Extract street name between ">" and "</option>""
						st = line[line.find(end_num)+1:line.find(end_name)]
						
						#Deal with Multiple TYPEs in a single line
						split_type_test = st.split(' ')[-1]
						for split_type in split_type_test.split('/') :
							st = ' '.join(st.split(' ')[:-1])+' '+split_type
							
							st = sm_standardize(st.strip())
							#Initialize dictionary for NAME if it doesn't exist already
							NAME = st[2]
							sm_st_ed_dict_city[NAME] = sm_st_ed_dict_city.setdefault(NAME, {})
							#Extract EDs as one long string
							st_str = line[line.find(start_num) + 7 : line.find(end_num)-1]
							#Check for street name collisions (e.g. "3rd Pl" and "3rd Pl ext")
							if st[0] in sm_st_ed_dict_city[NAME].keys():
								for i in list(st_str.split(",")):
									if i not in sm_st_ed_dict_city[NAME][st[0]]:
										sm_st_ed_dict_city[NAME][st[0]].append(i)
							else:    
								#Split string of EDs into a list and assign to street name in dictionary
								if decade == 1920 :
									sm_st_ed_dict_city[NAME].update({st[0]:list([x.split('[')[0] for x in st_str.split(",")])})
								else :
									sm_st_ed_dict_city[NAME].update({st[0]:list(st_str.split(","))})
		sm_st_ed_dict[decade][city_state] = sm_st_ed_dict_city
		time.sleep(0.5)
	print("Missing Steve Morse street-ED data for %s cities in %s" % (str(no_data),str(decade)))

# Output list usable by students
def output_usable_list(sm_st_ed_dict, decade, city_state_iterator):

	sm_dict = sm_st_ed_dict[decade]

	for city_state in city_state_iterator.itertuples(index=False):

		city = city_state[0].replace(' ','')
		state = city_state[1]
		c_s = (city,state)
		file_name = file_path + '/' + str(decade) + '/sm_lists/' + city + state.upper() + '_SM.xlsx'
		wb = Workbook()

		temp = sm_dict[c_s]
		sm_st_ed_dict_city = {k:v for d in [v for k,v in temp.items()] for k,v in d.items()}

		#Save street-to-ED sheet
		ws_st = wb.active
		ws_st.title = '%s Street to ED' % (str(decade)) 
		ws_st.append([str(decade) + ' Street',str(decade) + ' EDs'])
		dictlist = []
		keylist = sm_st_ed_dict_city.keys()
		keylist.sort()
		for k in keylist:
			t = [i.replace("'","") for i in sm_st_ed_dict_city[k]]
			ws_st.append([ignore_unicode(k)]+t)        

		#Save ED-to-street sheet
		ws_ed = wb.create_sheet('%s ED to Street' % (str(decade)))
		ws_ed.append([str(decade) + ' ED',str(decade) + ' Streets'])

		sm_ed_st_dict_city = {}
		for st, eds in sm_st_ed_dict_city.items():
			for ed in eds:
				#Initalize an empty list if ED has no list of street names yet
				sm_ed_st_dict_city[ed] = sm_ed_st_dict_city.setdefault(ed, [])
				#Add street name to the list of streets
				sm_ed_st_dict_city[ed].append(st)

		dictlist = []
		keylist = sm_ed_st_dict_city.keys()
		keylist.sort()
		for k in keylist:
			t = [ignore_unicode(i).replace("'","") for i in sm_ed_st_dict_city[k]]
			ws_ed.append([k]+t)       

		#If 1930, also include list of 1940 SM streets
		if decade == 1930:
			ws_st1940 = wb.create_sheet('1940 Streets')
			ws_st1940.append(['1940 Streets'])
			streetlist = sm_st_ed_dict[1940][c_s].keys()
			streetlist.sort()
			for i in streetlist:
				ws_st1940.append([i])

		wb.save(file_name)

# Scrape Steve Morse street-ed data from website
def scrape_sm_st_ed(file_path, decades=[1900,1910,1920,1930,1940]):

	city_info_file = file_path + '/CityExtractionList.csv' 
	city_info_df = pd.read_csv(city_info_file)
	city_info_df = city_info_df[city_info_df['Status']>0]
	city_info_df.loc[:,'city_name'], city_info_df.loc[:,'state_abbr'] = zip(*city_info_df['City'].str.split(','))
	city_info_df = city_info_df[['city_name','state_abbr']]
	city_info_df['city_name'] = city_info_df['city_name'].str.replace('Saint','St').str.replace('.','')
	city_info_df['state_abbr'] = city_info_df['state_abbr'].str.replace(' ','').str.lower()
	# Add New York boroughs
	new_york = [{'city_name':'Staten Island','state_abbr':'ny'}, 
		{'city_name':'Queens','state_abbr':'ny'}, 
		{'city_name':'Manhattan','state_abbr':'ny'}, 
		{'city_name':'Brooklyn','state_abbr':'ny'},
		{'city_name':'Bronx','state_abbr':'ny'}]
	city_info_df = city_info_df.append(pd.DataFrame(new_york))

	state_list = city_info_df['state_abbr'].tolist()

	global city_state_iterator
	city_state_iterator = city_info_df[['city_name','state_abbr']]

	# Download and save Steve Morse web abbreviations
	print("Downloading Steve Morse web abbreviations")
	sm_web_abbr_dict = {}
	for decade in decades:
		sm_web_abbr_dict[decade] = {}
		for state_abbr in state_list:
			sm_web_abbr_dict[decade][state_abbr] = {}
		get_sm_web_abbr(decade,sm_web_abbr_dict)
	print("Saving Steve Morse web abbreviations")
	pickle.dump(sm_web_abbr_dict,open(package_path + '/text/sm_web_abbr.pickle','wb'))

	# Download and save Steve Morse street-ed information
	print("Downloading Steve Morse street-ed information")
	sm_st_ed_dict ={}
	for decade in decades:
		sm_st_ed_dict[decade] = {}
		for i, row in city_state_iterator.iterrows():
			row = list(row)
			row[0] = row[0].replace(' ','')
			row = tuple(row)
			sm_st_ed_dict[decade][row] = {}
		get_sm_st_ed(decade, sm_st_ed_dict, package_path)
		#pickle.dump(sm_st_ed_dict[decade],open(file_path + '/%s/sm_st_ed_dict%s.pickle' % (str(decade),str(decade)),'wb'))
	print("Saving Steve Morse street-ed information")
	pickle.dump(sm_st_ed_dict,open(package_path + '/text/sm_st_ed_dict.pickle','wb'))

	# Output lists for use by students
	sm_st_ed_dict = pickle.load(open(package_path + '/text/sm_st_ed_dict.pickle','rb'))
	for decade in decades:
		#sm_ed_st_dict = create_ed_st_dict(sm_st_ed_dict[decade], city_state_iterator)
		output_usable_list(sm_st_ed_dict, decade, city_state_iterator)

# Load Steve Morse st_ed data
def load_steve_morse(city_info):

	city_name, state_abbr, decade = city_info

	#if city_name == 'Richmond' and state_abbr == 'NY':
	#	city_name = 'Staten Island'
	#NOTE: This dictionary must be built independently of this script
	package_path = os.path.dirname(histcensusgis.__file__)
	try :
		sm_st_ed_dict_file = pickle.load(open(package_path + '/text/sm_st_ed_dict.pickle', 'rb'))
	except ValueError :
		sm_st_ed_dict_file = pickle.load(open(package_path + '/text/sm_st_ed_dict.pickle', 'r'))
	sm_st_ed_dict_nested = sm_st_ed_dict_file[decade][(city_name.replace(' ',''), state_abbr.lower())]

	#Flatten dictionary
	temp = {k:v for d in [v for k, v in sm_st_ed_dict_nested.items()] for k, v in d.items()}

	#Capture all Steve Morse streets in one list
	sm_all_streets = temp.keys()

	#
	# Build a Steve Morse (sm) ED-to-Street (ed_st) dictionary (dict)
	#

	sm_ed_st_dict = {}
	#Initialize a list of street names without an ED in Steve Morse
	sm_ed_st_dict[''] = []
	for st, eds in temp.items():
		#If street name has no associated EDs (i.e. street name not found in Steve Morse) 
		#then add to dictionary entry for no ED
		if eds is None:
			sm_ed_st_dict[''].append(st)
		else:
			#For every ED associated with a street name...
			for ed in eds:
				#Initalize an empty list if ED has no list of street names yet
				sm_ed_st_dict[ed] = sm_ed_st_dict.setdefault(ed, [])
				#Add street name to the list of streets
				sm_ed_st_dict[ed].append(st)

	return sm_all_streets, sm_st_ed_dict_nested, sm_ed_st_dict


#download description data (called by get_sm_ed_desc)
def get_sm_descriptions(year):

	for city_state in city_state_iterator:
		print("getting info for "+str(city_state))
		sm_ed_descriptions[year][city_state] = {}

		city_name = city_state[0]
		state_abbr = city_state[1]

		url = "https://stevemorse.org/ed/%sdescriptions.%s.txt" % (str(year),state_abbr.upper()) 
		url_handle = urllib.urlopen(url)
		sourcetext = url_handle.readlines()
		url_handle.close()

		for line in sourcetext:
			line = line.lower()
			if city_name in line :
				if year == 1930 :
					ed_city_patt = "\^[0-9]+\-([0-9]+)\^"+re.escape(city_name)+"( city| borough)?[ /()nsew]*(,| \(part\)[ ,])"
					description_patt = "bounded by (.+)\n"
				if year == 1950 :
					ed_city_patt = "\^[0-9]+\-([0-9]+)\^"+re.escape(city_name)+"( city| borough)?[ /()nsew]*(,| \(part\)[ ,])"
					description_patt = "bounded by (.+)\n"
				ed_city = re.search(ed_city_patt,line)

				if ed_city :
					#line is a description of an ed in city
					ed = ed_city.group(1)
					description = re.search(description_patt,line)
					
					if description == None :
						#there is no description available for this city/ed
						sm_ed_descriptions[year][city_state][ed] = "description n/a"
					else :
						sm_ed_descriptions[year][city_state][ed] = description.group(1)


def format_descriptions(year,keep_dir) :
	for city_state in city_state_iterator:
		if year == 1930 :
			stphrase_patt = "\( ?([nesw\.]+) ?\) ?(.+)"
			for ed, desc in sm_ed_descriptions[year][city_state].items() :
				st_list = []
				for st in desc.split(";") :
					if st == ' (no population).' :
						pass
					stphrase_search = re.search(stphrase_patt,st)
					if stphrase_search :
						if keep_dir :
							st_list.append('('+stphrase_search.group(1).upper()+')')
						stphrase = stphrase_search.group(2)
						for stname in stphrase.split(", ") :
							st_list.append(sm_standardize(standardize_street(stname)[0])[0])
					elif st!="description n/a" :
						for stname in re.split(',|;',st) :
							st_list.append(sm_standardize(standardize_street(stname)[0])[0])
						
						print("ed "+ed+": problem with stname: "+st)
				
				sm_ed_descriptions[year][city_state][ed] = st_list
		elif year == 1950 :
			stphrase_patt = "(?:\([nesw]+\) )?(.+)"
			for ed, desc in sm_ed_descriptions[year][city_state].items() :
				st_list = []
				for st in desc.split(";") :
					st = st.strip()
					stphrase = re.search(stphrase_patt,st)
					if stphrase :
						stphrase = stphrase.group(1)
						for stname in stphrase.split(", ") :
							st_list.append(sm_standardize(standardize_street(stname)[0])[0])
					else :
						if st!="description n/a" :
							print("problem with stname: "+st)
				
				sm_ed_descriptions[year][city_state][ed] = st_list
		else :
			print (str(year)+" is not yet supported for descriptions.")

# Get the description data from Steve Morse for all cities in given year
# Unless otherwise specified, the ED directionals will be removed from the descriptions
# On the other hand, if keep_dir=True, the descriptions will not be formatted / standardized
def download_sm_ed_desc(year,keep_dir=False,file_path='S:/Projects/1940Census',output_path='S:/Projects/1940Census/SMdescriptions') :
	city_info_file = file_path + '/CityExtractionList.csv' 
	city_info_df = pd.read_csv(city_info_file)
	city_info_df = city_info_df[city_info_df['Status']>0]
	city_info_df.loc[:,'city_name'], city_info_df.loc[:,'state_abbr'] = zip(*city_info_df['City'].str.split(','))
	city_info_df = city_info_df[['city_name','state_abbr']]
	city_info_df['city_name'] = city_info_df['city_name'].str.lower().str.replace('saint','st.') #have to keep "St. ____"
	city_info_df['state_abbr'] = city_info_df['state_abbr'].str.replace(' ','').str.lower()
	# Add New York boroughs
	new_york = [{'city_name':'richmond','state_abbr':'ny'}, 
		{'city_name':'queens','state_abbr':'ny'}, 
		{'city_name':'manhattan','state_abbr':'ny'}, 
		{'city_name':'brooklyn','state_abbr':'ny'},
		{'city_name':'bronx','state_abbr':'ny'}]
	city_info_df = city_info_df.append(pd.DataFrame(new_york))
	global city_state_iterator
	city_state_iterator = zip(city_info_df['city_name'],city_info_df['state_abbr'])
	'''global city_state_iterator
	city_state_iterator = [('spokane',"wa"),
                               ('scranton',"pa"),
                               ('salt lake',"ut")]'''
	global sm_ed_descriptions
	sm_ed_descriptions = {}
	print("Starting "+str(year))
	sm_ed_descriptions[year] = {}
	
	get_sm_descriptions(year)
	
	format_descriptions(year,keep_dir)
	
	for city_state in city_state_iterator:
		file_name = output_path+"/"+city_state[0].title().replace(' ','')+city_state[1].upper()+"_SM_ED_desc.txt"
		file_path = open(file_name,"w+")

		for ed, desc in sorted(sm_ed_descriptions[year][city_state].items(),key=lambda x:int(re.search('[0-9]+',x[0]).group(0))) :
			string = ed+": "+", ".join(desc).replace("Railroad Tracks","Railway")
			if keep_dir :
				string = string.replace('),',')').replace(', (',' (')
			file_path.write(string+"\n")
		#print(ed+", "+str(desc).replace("[","").replace("]","").replace("Railroad Tracks","Railway"))
		#pass
		file_path.close()


### Functions for creating and updating the separate "full city" dictionary 
### that combines all unique street names form SM and the manually examined street grid (names only, not geometries)
# Note: eventually need a function to take streets added to grid and add to these dictionaries

# this function creates an empty pickle object to update later with full city street lists
def initialize_full_city_street_dict(file_path = '/home/s4-data/LatestCities'):

	city_info_file = file_path + '/CityExtractionList.csv' 
	city_info_df = pd.read_csv(city_info_file)
	city_info_df = city_info_df[city_info_df['Status']>0]
	city_info_df.loc[:,'city_name'], city_info_df.loc[:,'state_abbr'] = zip(*city_info_df['City'].str.split(','))
	city_info_df = city_info_df[['city_name','state_abbr']]
	city_info_df['city_name'] = city_info_df['city_name'].str.replace('Saint','St').str.replace('.','')
	city_info_df['state_abbr'] = city_info_df['state_abbr'].str.replace(' ','').str.lower()
	# Add New York boroughs
	new_york = [{'city_name':'Staten Island','state_abbr':'ny'}, 
		{'city_name':'Queens','state_abbr':'ny'}, 
		{'city_name':'Manhattan','state_abbr':'ny'}, 
		{'city_name':'Brooklyn','state_abbr':'ny'},
		{'city_name':'Bronx','state_abbr':'ny'}]
	city_info_df = city_info_df.append(pd.DataFrame(new_york))

	full_city = {a:None for a in zip(city_info_df['city_name'],city_info_df['state_abbr'])}

	pickle.dump(full_city, open(package_path + '/text/full_city_street_dict.pickle','wb'))



# this function combines all unique street names from grid and Steve Morse
# updates full_city_street_dict.pickle entry for city with full list
# You must upload 
def create_full_city_street_list(city_name, state_abbr, file_path = '/home/s4-data/LatestCities'):

	# load unique list of grid names (or return error that it doesn't exist)
	try:
		grid_dbf = gpd.read_file(file_path + '/grid_dbfs/' + city_name + state_abbr.upper() + '_stgrid.dbf')
	except IOError:
		sys.exit("No grid information found. Did you place this grid's .dbf file in '/home/s4-data/LatestCities/grid_dbfs'?")

	grid_streets = list(grid_dbf['st40'].unique())
	if grid_dbf['st30'].unique() != None:
		grid_streets.extend(list(grid_dbf['st30'].unique()))
		grid_streets = list(set(grid_streets))
	if grid_dbf['Alt_name'].unique() != None:
		grid_streets.extend(list(grid_dbf['st30'].unique()))
		grid_streets = list(set(grid_streets))

	# drop None values from list 
	grid_streets = [x for x in grid_streets if x]

	# convert unicode to str
	#grid_streets = map(str, grid_streets)
	grid_streets = [x.encode('ascii', 'replace') for x in grid_streets]

	# Load all avilable Steve Morse files for city
	sm_streets = []
	for decade in [1900, 1910, 1920, 1930, 1940]:
		try:
			all_streets, _, _ = load_steve_morse([city_name, state_abbr.lower(), decade])
			sm_streets.extend(all_streets)
			sm_streets = list(set(sm_streets))
		except:
			continue

	# combine lists and keep unique names
	sm_streets.extend(grid_streets)
	combined_streets = list(set(sm_streets))

	# open dictionary of full street lists
	try:
		full_city = pickle.load(open(package_path + '/text/full_city_street_dict.pickle', 'rb'))
	except ValueError:
		full_city = pickle.load(open(package_path + '/text/full_city_street_dict.pickle', 'r'))

	# update list for this city
	full_city[(city_name, state_abbr.lower())] = combined_streets

	# save updated dictionary
	pickle.dump(full_city, open(package_path + '/text/full_city_street_dict.pickle','wb'))

	# save street list as text file for manual street name cleaning
	os.chdir(file_path + '/full_street_lists')
	outfile = open(city_name + state_abbr.upper() + '_street_names.txt', 'w')
	for line in sorted(combined_streets):
		outfile.write(line + '\n')
	outfile.close()




