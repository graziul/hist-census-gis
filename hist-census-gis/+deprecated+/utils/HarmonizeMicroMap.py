from openpyxl import Workbook
from openpyxl import load_workbook
import shapefile, re
import fuzzyset
import pandas as pd
import pickle
from dbfpy import dbf

file_path = '/home/s4-data/LatestCities' 

city = 'Houston'
state = 'TX'
year = 1930

def standardize_street(st):
    runAgain = False
    st = st.rstrip('\n')
    orig_st = st
    
    st = st.lower()

    ###Remove Punctuation, extraneous words at end of stname###
    st = re.sub(r'[\.,]',' ',st)
    st = re.sub(' +',' ',st)
    st = st.strip()
    st = re.sub('\\\\','',st)
    st = re.sub(r' \(?([Cc][Oo][Nn][\'Tt]*d?|[Cc][Oo][Nn][Tt][Ii][Nn][Uu][Ee][Dd])\)?$','',st)
    #consider extended a diff stname#
    #st = re.sub(r' [Ee][XxsS][tdDT]+[^ ]*$','',st)

    #Check if st is empty or blank and return empty to [st,DIR,NAME,TYPE]
    if st == '' or st == ' ':
        return ['','','','']
    
    ###stname part analysis###
    DIR = ''
    NAME = ''
    TYPE = ''
 
    # Combinations of directions at end of stname (has to be run first)
    if re.search(r'[ \-]+([Nn][\.\-]?[\s]?[Ee][\.]?|[Nn]ortheast|[Nn]orth\s+?[Ee]ast)$',st):
        st = "NE "+re.sub(r'[ \-]+([Nn][\.\-]?[\s]?[Ee][\.]?|[Nn]ortheast|[Nn]orth[\s]+?[Ee]ast)$','',st)
        DIR = 'NE'
    if re.search(r'[ \-]+([Nn][\.\-]?[\s]?[Ww][\.]?|[Nn]orthwest|[Nn]orth\s+?[Ww]est)$',st):
        st = "NW "+re.sub(r'[ \-]+([Nn][\.\-]?[\s]?[Ww][\.]?|[Nn]orthwest|[Nn]orth\s+?[Ww]est)$','',st)
        DIR = 'NW'
    if re.search(r'[ \-]+([Ss][\.\-]?[\s]?[Ee][\.]?|[Ss]outheast|[Ss]outh\s+?[Ee]ast)$',st):
        st = "SE "+re.sub(r'[ \-]+([Ss][\.\-]?[\s]?[Ee][\.]?|[Ss]outheast|[Ss]outh\s+?[Ee]ast)$','',st)
        DIR = 'SE'
    if re.search(r'[ \-]+([Ss][\.\-]?[\s]?[Ww][\.]?|[Ss]outhwest|[Ss]outh\s+?[Ww]est)$',st):
        st = "SW "+re.sub(r'[ \-]+([Ss][\.\-]?[\s]?[Ww][\.]?|[Ss]outhwest|[Ss]outh\s+?[Ww]est)$','',st)
        DIR = 'SW'
   
    #First check if DIR is at end of stname. make sure that it's the DIR and not actually the NAME (e.g. "North Ave" or "Avenue E")#
    if re.search(r'[ \-]+([Nn]|[Nn][Oo][Rr]?[Tt]?[Hh]?)$',st) and not re.match('^[Nn][Oo][Rr][Tt][Hh]$|[Aa][Vv][Ee]([Nn][Uu][Ee])?[ \-]+[Nn]$',st) :
        st = "N "+re.sub(r'[ \-]+([Nn]|[Nn][Oo][Rr]?[Tt]?[Hh]?)$','',st)
        DIR = 'N'
    if re.search(r'[ \-]+([Ss]|[Ss][Oo][Uu]?[Tt]?[Hh]?)$',st) and not re.search('^[Ss][Oo][Uu][Tt][Hh]$|[Aa][Vv][Ee]([Nn][Uu][Ee])?[ \-]+[Ss]$',st) :
        st = "S "+re.sub(r'[ \-]+([Ss]|[Ss][Oo][Uu]?[Tt]?[Hh]?)$','',st)
        DIR = 'S'
    if re.search(r'[ \-]+([Ww][Ee][Ss][Tt]|[Ww])$',st) and not re.search('^[Ww][Ee][Ss][Tt]$|[Aa][Vv][Ee]([Nn][Uu][Ee])?[ \-]+[Ww]$',st) :
        st = "W "+re.sub(r'[ \-]+([Ww][Ee][Ss][Tt]|[Ww])$','',st)
        DIR = 'W'
    if re.search(r'[ \-]+([Ee][Aa][Ss][Tt]|[Ee])$',st) and not re.search('^[Ee][Aa][Ss][Tt]$|[Aa][Vv][Ee]([Nn][Uu][Ee])?[ \-]+[Ee]$',st) :
        st = "E "+re.sub(r'[ \-]+([Ee][Aa][Ss][Tt]|[Ee])$','',st)
        DIR = 'E'

    #See if a st TYPE can be identified#
    st = re.sub(r'[ \-]+([Ss][Tt][Rr]?[Ee]?[Ee]?[Tt]?[SsEe]?|[Ss][\.][Tt]|[Ss][Tt]?.?[Rr][Ee][Ee][Tt])$',' St',st)
    #st = re.sub(r'[ \-]+[Ss]tr?e?e?t?[ \-]',' St ',st) # Fix things like "4th Street Place"
    st = re.sub(r'[ \-]+([Aa][Vv]|[Aa][VvBb][Ee][Nn][Uu]?[EesS]?|[aA]veenue|[Aa]vn[e]?ue|[Aa][Vv][Ee])$',' Ave',st)
    match = re.search("[Aa][Vv][Ee]([Nn][Uu][Ee])?[ \-]+([a-zA-Z])$",st)
    if match :
         st = re.sub("([a-zA-Z])$","",st)
         st = re.sub("[Aa][Vv][Ee]([Nn][Uu][Ee])?[ \-]+",match.group(2)+" Ave",st)
    st = re.sub(r'[ \-]+([Bb]\'?[Ll][Vv]\'?[Dd]|Bl\'?v\'?d|Blv|Blvi|Bly|Bldv|Bvld|Bol\'d|[Bb][Oo][Uu][Ll][EeAa]?[Vv]?[Aa]?[Rr]?[Dd]?)$',' Blvd',st)
    st = re.sub(r'[ \-]+([Rr][Dd]|[Rr][Oo][Aa][Dd])$',' Road',st)
    st = re.sub(r'[ \-]+[Dd][Rr][Ii]?[Vv]?[Ee]?$',' Drive',st)
    st = re.sub(r'[ \-]+([Cc][Oo][Uu]?[Rr][Tt]|[Cc][Tt])$',' Ct',st)
    st = re.sub(r'[ \-]+([Pp][Ll][Aa]?[Cc]?[Ee]?)$',' Pl',st)
    st = re.sub(r'[ \-]+([Ss][Qq][Uu]?[Aa]?[Rr]?[Ee]?)$',' Sq',st)
    st = re.sub(r'[ \-]+[Cc]ircle$',' Cir',st)
    st = re.sub(r'[ \-]+([Pp]rkway|[Pp]arkway|[Pp]ark [Ww]ay|[Pp]kway|[Pp]ky|[Pp]arkwy|[Pp]rakway|[Pp]rkwy|[Pp]wy)$',' Pkwy',st)
    st = re.sub(r'[ \-]+[Ww][Aa][Yy]$',' Way',st)
    st = re.sub(r'[ \-]+[Aa][Ll][Ll]?[Ee]?[Yy]?$',' Aly',st)
    st = re.sub(r'[ \-]+[Tt][Ee][Rr]+[EeAa]?[Cc]?[Ee]?$',' Ter',st)
    st = re.sub(r'[ \-]+([Ll][Aa][Nn][Ee]|[Ll][Nn])$',' Ln',st)
    st = re.sub(r'[ \-]+([Pp]lzaz|[Pp][Ll][Aa][Zz][Aa])$',' Plaza',st)
    st = re.sub(r'[ \-]+([Hh]ighway)$',' Hwy',st)
    st = re.sub(r'[ \-]+([Hh]eights?)$',' Heights',st)

    # "Park" is not considered a valid TYPE because it should probably actually be part of NAME #
    match = re.search(r' (St|Ave|Blvd|Pl|Drive|Road|Ct|Railway|CityLimits|Hwy|Fwy|Pkwy|Cir|Ter|Ln|Way|Trail|Sq|Aly|Bridge|Bridgeway|Walk|Heights|Crescent|Creek|River|Line|Plaza|Esplanade|[Cc]emetery|Viaduct|Trafficway|Trfy|Turnpike)$',st)
    if match :
        TYPE = match.group(1)

    #Combinations of directions
    
    match = re.search(r'^([Nn][Oo\.]?[\s]?[Ee][\.]?|[Nn]ortheast|[Nn]orth\s+?[Ee]ast)[ \-]+',st)
    if match :
        if st == match.group(0)+TYPE :
            NAME = 'Northeast'
        else :
            st = "NE "+re.sub(r'^([Nn][Oo\.]?[\s]?[Ee][\.]?|[Nn]ortheast|[Nn]orth\s+?[Ee]ast)[ \-]+','',st)
            DIR = 'NE'
    match = re.search(r'^([Nn][Oo\.]?[\s]?[Ww][\.]?|[Nn]orthwest|[Nn]orth\s+?[Ww]est)[ \-]+',st)
    if match :
        if st == match.group(0)+TYPE :
            NAME = 'Northwest'
        else :
            st = "NW "+re.sub(r'^([Nn][Oo\.]?[\s]?[Ww][\.]?|[Nn]orthwest|[Nn]orth\s+?[Ww]est)[ \-]+','',st)
            DIR = 'NW'
    match = re.search(r'^([Ss][Oo\.]?[\s]?[Ee][\.]?|[Ss]outheast|[Ss]outh\s+?[Ee]ast)[ \-]+',st)
    if match :
        if st == match.group(0)+TYPE :
            NAME = 'Southeast'
        else :
            st = "SE "+re.sub(r'^([Ss][Oo\.]?[\s]?[Ee][\.]?|[Ss]outheast|[Ss]outh\s+?[Ee]ast)[ \-]+','',st)
            DIR = 'SE'
    match = re.search(r'^([Ss][Oo\.]?[\s]?[Ww][\.]?|[Ss]outhwest|[Ss]outh\s+?[Ww]est)[ \-]+',st)
    if match :
        if st == match.group(0)+TYPE :
            NAME = 'Southwest'
        else :
            st = "SW "+re.sub(r'^([Ss][Oo\.]?[\s]?[Ww][\.]?|[Ss]outhwest|[Ss]outh\s+?[Ww]est)[ \-]+','',st)
            DIR = 'SW'
        
    #See if there is a st DIR. again, make sure that it's the DIR and not actually the NAME (e.g. North Ave, E St [not East St])
    if(DIR=='') :
        match =  re.search(r'^([nN]|[Nn]\.|[Nn]o|[nN]o\.|[Nn][Oo][Rr][Tt]?[Hh]?)[ \-]+',st)
        if match :
            if st==match.group(0)+TYPE :
                if len(match.group(1))>1 :
                    NAME = 'North'
                else : NAME = 'N'
            else :
                st = "N "+re.sub(r'^([nN]|[Nn]\.|[Nn]o|[nN]o\.|[Nn][Oo][Rr][Tt]?[Hh]?)[ \-]+','',st)
                DIR = 'N'
        match =  re.search(r'^([sS]|[Ss]\.|[Ss]o|[Ss]o\.|[Ss][Oo][Uu][Tt]?[Hh]?)[ \-]+',st)
        if match :
            if st==match.group(0)+TYPE :
                if len(match.group(1))>1:
                    NAME = 'South'
                else : NAME = 'S'
            else :
                st = "S "+re.sub(r'^([sS]|[Ss]\.|[Ss]o|[Ss]o\.|[Ss][Oo][Uu][Tt]?[Hh]?)[ \-]+','',st)
                DIR = 'S'
        match =  re.search(r'^([wW]|[Ww]\.|[Ww][Ee][Ss]?[Tt]?[\.]?)[ \-]+',st)
        if match :
            if st==match.group(0)+TYPE :
                if len(match.group(1))>1 :
                    NAME = 'West'
                else : NAME = 'W'
            else :
                st = "W "+re.sub(r'^([wW]|[Ww]\.|[Ww][Ee][Ss]?[Tt]?[\.]?)[ \-]+','',st)
                DIR = 'W'
        match =  re.search(r'^([eE]|[Ee][\.\,]|[Ee][Ee]?[Aa]?[Ss][Tt][\.]?|[Ee]a[Ss]?)[ \-]+',st)
        if match :
            if st==match.group(0)+TYPE :
                if len(match.group(1))>1 :
                    NAME = 'East'
                else : NAME = 'E'
            else :
                st = "E "+re.sub(r'^([eE]|[Ee][\.\,]|[Ee][Ee]?[Aa]?[Ss][Tt][\.]?|[Ee]a[Ss]?)[ \-]+','',st)
                DIR = 'E'
                
    #get the st NAME and standardize it
            
    match = re.search('^'+DIR+'(.+)'+TYPE+'$',st)
    if NAME=='' :
        #If NAME is not 'North', 'West', etc...
        if match :
            NAME = match.group(1).strip()
            
            #convert written-out numbers to digits
            #TODO: Make these work for all exceptions (go thru text file with find)
            #if re.search("[Tt]enth|Eleven(th)?|[Tt]wel[f]?th|[Tt]hirteen(th)?|Fourt[h]?een(th)?|[Ff]ift[h]?een(th)?|[Ss]event[h]?een(th)?|[Ss]event[h]?een(th)?|[eE]ighteen(th)?|[Nn]inet[h]?een(th)?|[Tt]wentieth|[Tt]hirtieth|[Ff]o[u]?rtieth|[Ff]iftieth|[Ss]ixtieth|[Ss]eventieth|[Ee]ightieth|[Nn]inetieth|Twenty[ \-]?|Thirty[ \-]?|Forty[ \-]?|Fifty[ \-]?|Sixty[ \-]?|Seventy[ \-]?|Eighty[ \-]?|Ninety[ \-]?|[Ff]irst|[Ss]econd|[Tt]hird|[Ff]ourth|[Ff]ifth|[Ss]ixth|[Ss]eventh|[Ee]ighth|[Nn]inth",st) :
            NAME = re.sub("^[Tt]enth","10th",NAME)
            NAME = re.sub("^[Ee]leven(th)?","11th",NAME)
            NAME = re.sub("^[Tt]wel[f]?th","12th",NAME)
            NAME = re.sub("^[Tt]hirteen(th)?","13th",NAME)
            NAME = re.sub("^[Ff]ourt[h]?een(th)?","14th",NAME)
            NAME = re.sub("^[Ff]ift[h]?een(th)?","15th",NAME)
            NAME = re.sub("^[Ss]ixt[h]?een(th)?","16th",NAME)
            NAME = re.sub("^[Ss]event[h]?een(th)?","17th",NAME)
            NAME = re.sub("^[eE]ighteen(th)?","18th",NAME)
            NAME = re.sub("^[Nn]inet[h]?e+n(th)?","19th",NAME)
            NAME = re.sub("^[Tt]went[iy]eth","20th",NAME)
            NAME = re.sub("^[Tt]hirt[iy]eth","30th",NAME)
            NAME = re.sub("^[Ff]o[u]?rt[iy]eth","40th",NAME)
            NAME = re.sub("^[Ff]ift[iy]eth", "50th",NAME)
            NAME = re.sub("^[Ss]ixt[iy]eth", "60th",NAME)
            NAME = re.sub("^[Ss]event[iy]eth", "70th",NAME)
            NAME = re.sub("^[Ee]ight[iy]eth", "80th",NAME)
            NAME = re.sub("^[Nn]inet[iy]eth", "90th",NAME)

            NAME = re.sub("[Tt]wenty[ \-]*","2",NAME)
            NAME = re.sub("[Tt]hirty[ \-]*","3",NAME)
            NAME = re.sub("[Ff]orty[ \-]*","4",NAME)
            NAME = re.sub("[Ff]ifty[ \-]*","5",NAME)
            NAME = re.sub("[Ss]ixty[ \-]*","6",NAME)
            NAME = re.sub("[Ss]eventy[ \-]*","7",NAME)
            NAME = re.sub("[Ee]ighty[ \-]*","8",NAME)
            NAME = re.sub("[Nn]inety[ \-]*","9",NAME)
            
            if re.search("(^|[0-9]+.*)([Ff]irst|[Oo]ne)$",NAME) : NAME = re.sub("([Ff]irst|[Oo]ne)$","1st",NAME)
            if re.search("(^|[0-9]+.*)([Ss]econd|[Tt]wo)$",NAME) : NAME = re.sub("([Ss]econd|[Tt]wo)$","2nd",NAME)
            if re.search("(^|[0-9]+.*)([Tt]hird|[Tt]hree)$",NAME) : NAME = re.sub("([Tt]hird|[Tt]hree)$","3rd",NAME)
            if re.search("(^|[0-9]+.*)[Ff]our(th)?$",NAME) : NAME = re.sub("[Ff]our(th)?$","4th",NAME)
            if re.search("(^|[0-9]+.*)([Ff]ifth|[Ff]ive)$",NAME) : NAME = re.sub("([Ff]ifth|[Ff]ive)$","5th",NAME)
            if re.search("(^|[0-9]+.*)[Ss]ix(th)?$",NAME) : NAME = re.sub("[Ss]ix(th)?$","6th",NAME)
            if re.search("(^|[0-9]+.*)[Ss]even(th)?$",NAME) : NAME = re.sub("[Ss]even(th)?$","7th",NAME)
            if re.search("(^|[0-9]+.*)[Ee]igh?th?$",NAME) : NAME = re.sub("[Ee]igh?th?$","8th",NAME)
            if re.search("(^|[0-9]+.*)[Nn]in(th|e)+$",NAME) : NAME = re.sub("[Nn]in(th|e)+$","9th",NAME)
            
            if re.search("[0-9]+",NAME) :
                if re.search("^[0-9]+$",NAME) : #if NAME is only numbers (no suffix), add the correct suffix
                    foo = True
                    suffixes = {'11':'11th','12':'12th','13':'13th','1':'1st','2':'2nd','3':'3rd','4':'4th','5':'5th','6':'6th','7':'7th','8':'8th','9':'9th','0':'0th'}
                    num = re.search("[0-9]+$",NAME).group(0)
                    suff = ''
                    # if num is not found in suffixes dict, remove leftmost digit until it is found... 113 -> 13 -> 13th;    24 -> 4 -> 4th
                    while(suff=='') :
                        try :
                            suff = suffixes[num]
                        except KeyError :
                            num = num[1:]
                            if len(num) == 0 :
                                break
                    if not suff == '' :
                        NAME = re.sub(num+'$',suff,NAME)
                else :
                    # Fix incorrect suffixes e.g. "73d St" -> "73rd St"
                    if re.search("[23]d$",NAME) :
                        NAME = re.sub("3d","3rd",NAME)
                        NAME = re.sub("2d","2nd",NAME)
                    if re.search("1 [Ss]t|2 nd|3 rd|1[1-3] th|[04-9] th",NAME) :
                        try :
                            suff = re.search("[0-9] ([Sa-z][a-z])",NAME).group(1)
                        except :
                            print("NAME: "+NAME+", suff: "+suff+", st: "+st)
                        NAME = re.sub(" "+suff,suff,NAME)
                    # TODO: identify corner cases with numbers e.g. "51 and S- Hermit"
                
                # This \/ is a bit overzealous...! #
                hnum = re.search("^([0-9]+[ \-]+).+",NAME) #housenum in stname?
                if hnum : 
                    #False
                    NAME = re.sub(hnum.group(1),"",NAME) #remove housenum. May want to update housenum field, maybe not though.
                    runAgain = True
            else :
                NAME = NAME.title()
            
        else :
            assert(False)
        # Standardize "St ____ Ave" -> "Saint ____ Ave" #
        NAME = re.sub("^([Ss][Tt]\.?|[Ss][Aa][Ii][Nn][Tt])[ \-]","Saint ",NAME)
    st = re.sub(re.escape(match.group(1).strip()),NAME,st).strip()
    try :
        assert st == (DIR+' '+NAME+' '+TYPE).strip()
    except AssertionError :
        print("Something went a bit wrong while trying to pre-standardize stnames.")
        print("orig was: "+orig_st)
        print("st is: \""+st+"\"")
        print("components: ["+(DIR+','+NAME+','+TYPE).strip()+"]")
    
    if runAgain :
        return standardize_street(st)
    else :
        return [st, DIR, NAME, TYPE]

#
# Step 0: Pull in street lists
#

	# From map

fileShp = "/home/s4-data/LatestCities/1930/autostudcleaned/Houston_1930_stgrid_edit.shp"
sf = shapefile.Reader(fileShp)
st_shp = [r.record[0] for r in sf.shapeRecords()]

'''
fileDbf = "/home/s4-data/LatestCities/1930/autostudcleaned/Houston_1930_stgrid_edit.dbf"
db = dbf.Dbf(fileDbf)
st_shp = [r[0] for r in db]
'''

# Get unique list of streets in shapefile
st_shp = list(set(st_shp))

# Build {NAME:TYPE} dictionary
st_shp_dict = {}
st_shp_list = []
for st in st_shp:
	extracted = standardize_street(st)
	st_shp_list.append(extracted)
	st_shp_dict[extracted[2]] = st_shp_dict.setdefault(extracted[2], [])
	st_shp_dict[extracted[2]].append(extracted[3])
	st_shp_dict[extracted[2]] = list(set(st_shp_dict[extracted[2]]))

	# From microdata

fileDta = "/home/s4-data/LatestCities/1930/autostudcleaned/HoustonTX_StudAuto.dta"
df = pd.read_stata(fileDta)
st_micro = list(set(df['autostud_street']))

st_micro_dict = {}
st_micro_list = []
for st in st_micro:
	extracted = standardize_street(st)
	st_micro_list.append(extracted)
	st_micro_dict[extracted[2]] = st_micro_dict.setdefault(extracted[2], [])
	st_micro_dict[extracted[2]].append(extracted[1:])

'''
#
# Step 1: Do fuzzy matching for NAME (fix map names using microdata names)
#

def find_fuzzy_matches(names_micro,names_shp,check_too_similar=False):

	print("Fuzzy matching map names to match microdata names")

	#Find exact matches and exclude from fuzzy matching
	exact_matches = set(names_micro).intersection(names_shp)
	num_exact_matches = len(exact_matches)
	prop_exact_matches = float(num_exact_matches)/len(names_shp)
	print("\nExact matches: "+str(num_exact_matches)+" of "+str(len(names_shp))+" map NAMEs ("+str(round(100*prop_exact_matches,1))+"%)")

	#Create a set of all streets for fuzzy matching (create once, call on)
	names_micro_fuzzyset = fuzzyset.FuzzySet(names_micro)

    #Fuzzy matching algorithm
	def fuzzy_match(name):

		#Return null if DIRNAME is blank
		if name == '':
			return ['','']

        #Step 1: Find best match among all streets
		try:
			best_match_all = names_micro_fuzzyset[name][0]
		except:
			return ['','']    

        #Step 2: Make sure it's a very good match
		if best_match_all[0] >= 0.9:
			return [best_match_all[1],best_match_all[0]]
		else:
			return ['', '']
   	
   	names_shp_to_check = [i for i in names_shp if i not in exact_matches]
	fuzzy_match_dict = {}
	num_fuzzy_matches = 0
	for name in names_shp_to_check:
		potential_match = fuzzy_match(name)
		if potential_match[0] != '':
			fuzzy_match_dict[name] = potential_match[0]
			num_fuzzy_matches += 1
		else:
			continue

    #Generate fuzzy_info
	prop_fuzzy_matches = float(num_fuzzy_matches)/len(names_shp)
	print("Fuzzy matches: "+str(num_fuzzy_matches)+" of "+str(len(names_shp))+" map NAMEs ("+str(round(100*prop_fuzzy_matches,1))+"%)")

	num_unmatched = len(names_shp) - (num_exact_matches + num_fuzzy_matches)
	prop_unmatched = float(num_unmatched)/len(names_shp)
	print("Unmatched: "+str(num_unmatched)+" of "+str(len(names_shp))+" map NAMEs ("+str(round(100*prop_unmatched,1))+"%)")

	return fuzzy_match_dict

# Get list of NAMEs in shapefile
names_shp = st_shp_dict.keys()

# Get list of NAMEs in microdata
names_micro = st_micro_dict.keys()

# Build fuzzy match dictionary
fuzzy_match_dict = find_fuzzy_matches(names_micro,names_shp)

# Discard changes when NAME from map matches Steve Morse name
def load_steve_morse(city,state,year):

	#NOTE: This dictionary must be built independently of this script
	sm_st_ed_dict_file = pickle.load(open(file_path + '/%s/sm_st_ed_dict%s.pickle' % (str(year),str(year)),'rb'))
	sm_st_ed_dict_nested = sm_st_ed_dict_file[(city,'%s' % (state.lower()))]

	#Flatten dictionary
	temp = {k:v for d in [v for k,v in sm_st_ed_dict_nested.items()] for k,v in d.items()}

	#Capture all Steve Morse streets in one list
	sm_streets = temp.keys()

	return sm_streets
sm_streets = load_steve_morse(city,state,year)

st_SM_dict = {}
for st in sm_streets:
	extracted = standardize_street(st)
	st_SM_dict[extracted[2]] = st_SM_dict.setdefault(extracted[2], [])
	st_SM_dict[extracted[2]].append(extracted[1:])
sm_names = st_SM_dict.keys()

fuzzy_match_dict = {k:v for k,v in fuzzy_match_dict.items() if k not in sm_names}

# Replace NAMEs from shapefile using dictionary
for r in db:
	st_shp = r['FULLNAME']
	_,_,NAME_shp,_ = standardize_street(st_shp)
	if NAME_shp in fuzzy_match_dict.keys():
		r['FULLNAME'] = st_shp.replace(NAME_shp,fuzzy_match_dict[NAME_shp])
		r.store()
db.close()
'''

#
# Step 2: Determine which NAMEs have more than one TYPE
#

# Get list of NAMEs in shapefile with more than one TYPE
names_w_multiple_types_shp = [k for k,v in st_shp_dict.items() if len(v) > 1]

# Get list of NAMEs in microdata with more than one TYPE
names_w_multiple_types_micro = [k for k,v in st_micro_dict.items() if len(v) > 1]

# Combine lists of NAMEs with more than one TYPE
names_w_multiple_types = list(set(names_w_multiple_types_micro + names_w_multiple_types_shp))

#
# Step 3: Determine which NAMEs in microdata have no TYPE and build change dictionary

# Get list of NAMEs in microdata with no TYPE
names_w_no_type = [k for k,v in st_micro_dict.items() if (len(v) == 1) and (v[0][2] == '')]

add_type_dict = {}
for k,v in st_micro_dict.items():
	st = v[0]
	DIR = st[0]
	NAME = st[1]
	TYPE = st[2]
	if NAME in names_w_no_type and NAME not in names_w_multiple_types:
		if TYPE != '':
			print("Error, has type already")
			break
		if DIR == '':
			old_st = NAME
			new_st = NAME + ' St'
			add_type_dict[old_st] = new_st
		else:
			old_st = DIR + ' ' + NAME
			new_st = DIR + ' ' + NAME + ' St'
			add_type_dict[old_st] = new_st			

#
# Step 4: Update microdata 
#

def update_microdata(st):
	try:
		st_add_st = add_type_dict[st]
		return st_add_st
	except: 
		return st

df['autostud_street'] = df.apply(lambda x: update_microdata(x['autostud_street']),axis=1)	

#
# Step 5: Find streets in microdata but not map and save
#

st_in_microdata = df['autostud_street'].unique().tolist()
st_in_map = st_shp

st_in_microdata_not_map = [st for st in st_in_microdata if st not in st_in_map]
st_in_microdata_not_map.sort()
st_in_microdata_not_map.remove('.') 

ed_st_groups = df.groupby(['ed','autostud_street'])
ed_st_list = []
for name, _ in ed_st_groups:
	ed_st_list.append([name[1],name[0]])

st_to_check = []
for name in st_in_microdata_not_map:
	temp = [name]
	streets = [i[1] for i in ed_st_list if i[0] == name]
	for j in streets:
		temp.append(j)
	st_to_check.append(temp)

# Open Excel workbook
c_s = city + str.upper(state)
file_name = file_path + '/' + str(year) + '/autostudcleaned/' + c_s + '_SM.xlsx'
wb = load_workbook(file_name)

#Save streets to check on map
ws_st = wb.create_sheet('Streets to check on map')
ws_st.append(['Street','EDs where located'])
for i in st_to_check:
	ws_st.append(i)

wb.save(file_name)        

#
# Step 6: Note streets in microdata that are not in Steve Morse 
#

def load_steve_morse(city,state,year):

	#NOTE: This dictionary must be built independently of this script
	sm_st_ed_dict_file = pickle.load(open(file_path + '/%s/sm_st_ed_dict%s.pickle' % (str(year),str(year)),'rb'))
	sm_st_ed_dict_nested = sm_st_ed_dict_file[(city,'%s' % (state.lower()))]

	#Flatten dictionary
	temp = {k:v for d in [v for k,v in sm_st_ed_dict_nested.items()] for k,v in d.items()}

	#Capture all Steve Morse streets in one list
	sm_streets = temp.keys()

	return sm_streets
sm_streets = load_steve_morse(city,state,year)

st_not_in_SM = [i for i in st_in_microdata_not_map if i not in sm_streets]

# Open Excel workbook
c_s = city + str.upper(state)
file_name = file_path + '/' + str(year) + '/autostudcleaned/' + c_s + '_SM.xlsx'
wb = load_workbook(file_name)

#Save streets to check on map
ws_st = wb.create_sheet('Streets not in Steve Morse')
ws_st.append(['Street'])
for name in st_not_in_SM:
	ws_st.append([name])

wb.save(file_name) 