from openpyxl import Workbook
import urllib
import pandas as pd
import re
import pickle

# Use one year's CityInfo file to get list of states
file_path = '/home/s4-data/LatestCities'
city_info_file = file_path + '/CityInfo.csv' 
city_info_df = pd.read_csv(city_info_file)

state_list = city_info_df['state_abbr'].str.lower().unique().tolist()
city_list = city_info_df['city_name'].tolist()

city_state_iterator = zip(city_info_df['city_name'].tolist(),city_info_df['state_abbr'].str.lower().tolist())
    
# Get the city_state abbreviation for Steve Morse    
def get_sm_web_abbr(year):

    for city_state in city_state_iterator:

        city_name = city_state[0]
        state_abbr = city_state[1]

        sm_web_abbr_dict[year][state_abbr][city_name] = []

        url = "http://stevemorse.org/census/%sstates/%s.htm" % (str(year),state_abbr) 
        url_handle = urllib.urlopen(url)
        sourcetext = url_handle.readlines()
        url_handle.close()

        #Change the parsing strings if necessary (CG: Doesn't seem necessary yet)
        start_num = "value=" 
        end_num = ">"
        end_name = "</option>"

        for line in sourcetext:
            #If line has "value=" and "</option>" in it, do stuff
            if start_num in line and end_name in line:
                #Ignore header line
                if "Select City" in line:
                    continue
                if "Other" in line:
                    continue
                else:
                    city = line[line.find(end_num)+1:line.find("(")].rstrip().replace('.','')
                    if city_name in city:
                        city_abbr = line[line.find(start_num) + 7 : line.find(end_num)-1].lower()
                        city_abbr = re.sub(r'\$[0-9]|\*[0-9]','',city_abbr)
                        sm_web_abbr = city_abbr + state_abbr
                        if sm_web_abbr not in sm_web_abbr_dict[year][state_abbr][city_name]:
                            sm_web_abbr_dict[year][state_abbr][city_name].append(sm_web_abbr)


sm_web_abbr_dict = {}
for year in [1900,1910,1930,1940]:
    sm_web_abbr_dict[year] = {}
    for state_abbr in state_list:
        sm_web_abbr_dict[year][state_abbr] = {}
    get_sm_web_abbr(year)

pickle.dump(sm_web_abbr_dict,open(file_path + '/sm_web_abbr.pickle','wb'))

#Standardize Steve Morse street names
def sm_standardize(st) :
    orig_st = st
    st = re.sub(r" [Ee][Xx][Tt][Ee]?[Nn]?[Dd]?[Ee]?[Dd]?$","",st)
    DIR = re.search(r" ([NSEW ]+)$",st)
    st = re.sub(r" ([NSEW ]+)$","",st)
    if(DIR) :
        DIR = DIR.group(1)
        DIR = re.sub(" ","",DIR)
    else :
        DIR = ""

    TYPE = re.search(r' (St|Ave?|Blvd|Pl|Dr|Drive|Rd|Road|Ct|Railway|CityLimits|Hwy|Fwy|Pkwy|Cir|Ter|Ln|Way|Trail|Sq|All?e?y?|Bridge|Bridgeway|Walk|Crescent|Creek|River|Line|Plaza|Esplanade|[Cc]emetery|Viaduct|Trafficway|Trfy|Turnpike)$',st)
    
    if(TYPE) :
        st = re.sub(TYPE.group(0),"",st)
        TYPE = TYPE.group(1)
        if(TYPE=="Av") :
            TYPE = "Ave"
        if(TYPE=="Rd") :
            TYPE = "Road"
        if(TYPE=="Dr") :
            TYPE = "Drive"
        if(re.match("All?e?y?",TYPE)) :
            TYPE = "Aly"
    else :
        TYPE = "St"
    
    NAME = st
    st = (DIR+" "+NAME+" "+TYPE).strip()
    #print(orig_st)
    #print("changed to "+st)
    return [st,DIR,NAME,TYPE]

# Get Steve Morse street-ed information
def get_sm_st_ed(year):
        
    for city_state in city_state_iterator:

        city_name = city_state[0]
        state_abbr = city_state[1]
  
        sm_st_ed_dict_city = {}

        try:
            sm_web_abbr = sm_web_abbr_dict[year][state_abbr][city_name]
        except:
            continue

        for i in sm_web_abbr:

            url = "http://www.stevemorse.org/census/%scities/%s.htm" % (str(year),i) 

            url_handle = urllib.urlopen(url)
            sourcetext = url_handle.readlines()
            url_handle.close()

            #Change the parsing strings if necessary (CG: Doesn't seem necessary yet)
            start_num = "value=" 
            end_num = ">"
            end_name = "</option>"

            for line in sourcetext:
                #If line has "value=" and "</option>" in it, do stuff
                if start_num in line and end_name in line:
                    #Ignore header line
                    if "Select Street" in line:
                        col_line = ""
                    else:
                        #Extract street name between ">" and "</option>""
                        st = line[line.find(end_num)+1:line.find(end_name)]
                        st = sm_standardize(st)
                        #Initialize dictionary for NAME if it doesn't exist already
                        NAME = st[2]
                        sm_st_ed_dict_city[NAME] = sm_st_ed_dict_city.setdefault(NAME, {})
                        #Extract EDs as one long string
                        st_str = line[line.find(start_num) + 7 : line.find(end_num)-1]
                        #Check for street name collisions (e.g. "3rd Pl" and "3rd Pl ext")
                        if st[0] in sm_st_ed_dict_city[NAME].keys():
                            for i in list(st_str.split(",")):
                                if i not in sm_st_ed_dict_city[NAME][st[0]]:
                                    sm_st_ed_dict_city[NAME][st[0]].append(i)
                        else:    
                            #Split string of EDs into a list and assign to street name in dictionary
                            sm_st_ed_dict_city[NAME].update({st[0]:list(st_str.split(","))})
        sm_st_ed_dict[year][city_state] = sm_st_ed_dict_city

sm_st_ed_dict ={}
for year in [1900,1910,1930,1940]:
    sm_st_ed_dict[year] = {}
    for city_state in city_state_iterator:
        sm_st_ed_dict[year][city_state] = {}
    get_sm_st_ed(year)

for year in [1900,1910,1930,1940]:
    pickle.dump(sm_st_ed_dict[year],open(file_path + '/%s/sm_st_ed_dict%s.pickle' % (str(year),str(year)),'wb'))

# Create usable ST-ED and ED-ST lists

sm_ed_st_dict = {}
for year in [1900,1910,1930,1940]:
    sm_ed_st_dict[year] = {}
    for city_state in city_state_iterator:
        sm_ed_st_dict[year][city_state] = {}
        sm_st_ed_dict_nested = sm_st_ed_dict[year][city_state] 
        #Flatten dictionary
        temp = {k:v for d in [v for k,v in sm_st_ed_dict_nested.items()] for k,v in d.items()}
        #Initialize a list of street names without an ED in Steve Morse
        sm_ed_st_dict[year][city_state][''] = []
        for st, eds in temp.items():
            #If street name has no associated EDs (i.e. street name not found in Steve Morse) 
            #then add to dictionary entry for no ED
            if eds is None:
                sm_ed_st_dict[year][city_state][''].append(st)
            else:
                #For every ED associated with a street name...
                for ed in eds:
                    #Initalize an empty list if ED has no list of street names yet
                    sm_ed_st_dict[year][city_state][ed] = sm_ed_st_dict[year][city_state].setdefault(ed, [])
                    #Add street name to the list of streets
                    sm_ed_st_dict[year][city_state][ed].append(st)

def output_usable_list(sm_st_ed_dict,year):

    sm_dict = sm_st_ed_dict[year]

    for city_state in city_state_iterator:
        c_s = city_state[0] + city_state[1].upper()
        file_name = file_path + '/' + str(year) + '/sm_lists/' + c_s + '_SM.xlsx'
        wb = Workbook()

        temp = sm_dict[city_state]
        sm_st_ed_dict_city = {k:v for d in [v for k,v in temp.items()] for k,v in d.items()}

        #Save street-to-ED sheet
        ws_st = wb.active
        ws_st.title = '1930 Street to ED' 
        ws_st.append([str(year) + ' Street',str(year) + ' EDs'])
        dictlist = []
        keylist = sm_st_ed_dict_city.keys()
        keylist.sort()
        for k in keylist:
            t = [i.replace("'","") for i in sm_st_ed_dict_city[k]]
            ws_st.append([k]+t)        

        #Save ED-to-street sheet
        ws_ed = wb.create_sheet('1930 ED to Street')
        ws_ed.append([str(year) + ' ED',str(year) + ' Streets'])

        sm_ed_st_dict_city = {}
        for st, eds in sm_st_ed_dict_city.items():
            for ed in eds:
                #Initalize an empty list if ED has no list of street names yet
                sm_ed_st_dict_city[ed] = sm_ed_st_dict_city.setdefault(ed, [])
                #Add street name to the list of streets
                sm_ed_st_dict_city[ed].append(st)

        dictlist = []
        keylist = sm_ed_st_dict_city.keys()
        keylist.sort()
        for k in keylist:
            t = [i.replace("'","") for i in sm_ed_st_dict_city[k]]
            ws_ed.append([k]+t)       

        #If 1930, also include list of 1940 SM streets
        if year == 1930:
            ws_st1940 = wb.create_sheet('1940 Streets')
            ws_st1940.append(['1940 Streets'])
            streetlist = sm_st_ed_dict[1940][city_state].keys()
            streetlist.sort()
            for s in streetlist:
                ws_st1940.append([s])

        wb.save(file_name)

'''
                col = 1         
                for i in sm_dict_city[k]:
                    row.write(k, col, i.replace("'",""))
                    col += 1    
'''

for year in [1900,1910,1930,1940]:
    output_usable_list(sm_st_ed_dict,year)
